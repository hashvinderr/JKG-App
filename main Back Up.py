import sys
import os
import json
import calendar
import traceback
from datetime import datetime, timedelta
from PyQt5 import QtWidgets, QtCore, QtGui

# Default skill for legacy assignments
DEFAULT_SKILL = "General Worker"
# Master list of skills for site assignments
SKILLS = ["General Worker", "Rigger/Signalman", "Lifting Supervisor", "Welder"]

# Sites configuration
DEFAULT_SITES = {
    "Asia Piling": {"lunch": 30, "extra_dinner": False},
    "CCECC": {"lunch": 60, "extra_dinner": True},
    "GS ENC N101": {"lunch": 60, "extra_dinner": False},
    "GS ENC T301": {"lunch": 60, "extra_dinner": False},
    "MCCONNELL DOWELL J108": {"lunch": 60, "extra_dinner": False},
    "NISHIMATSU CR110": {"lunch": 60, "extra_dinner": False},
    "NISHIMATSU CR210": {"lunch": 60, "extra_dinner": False},
    "PINTARY FOUNDATIONS": {"lunch": 60, "extra_dinner": False},
    "PROGRESS PILING": {"lunch": 30, "extra_dinner": False},
    "SATO KOGYO": {"lunch": 60, "extra_dinner": False},
    "SHANGHAI TUNNEL": {"lunch": 60, "extra_dinner": False},
    "WOH HUP": {"lunch": 60, "extra_dinner": False},
    "L&M": {"lunch": 60, "extra_dinner": False},
    "SAMWOH": {"lunch": 60, "extra_dinner": False},
    "SZ CONSTRUCTION": {"lunch": 60, "extra_dinner": False},
    "MK TUNNELS": {"lunch": 60, "extra_dinner": False},
    "METROCON PTE LTD": {"lunch": 60, "extra_dinner": False},
    "GREENMARK CONSTRUCTION": {"lunch": 60, "extra_dinner": False},
}

# Global exception handler to show errors instead of silent crashes
def get_config_path():
    """Return the path to the config.json in OneDrive or home folder."""
    od = os.environ.get("OneDriveCommercial") or os.environ.get("OneDrive")
    if od and os.path.isdir(od) and os.access(od, os.W_OK):
        base = od
    else:
        base = os.path.expanduser("~")
    folder = os.path.join(base, "JKGTimecardApp")
    os.makedirs(folder, exist_ok=True)
    return os.path.join(folder, "config.json")

# Global exception handler to show errors instead of silent crashes
def excepthook(exc_type, exc_value, exc_tb):
    msg = QtWidgets.QMessageBox(QtWidgets.QMessageBox.Critical, "Unhandled Exception", f"{exc_type.__name__}: {exc_value}")
    msg.setDetailedText(''.join(traceback.format_exception(exc_type, exc_value, exc_tb)))
    msg.exec_()
    sys.exit(1)
sys.excepthook = excepthook

# Configuration Handling
def load_config():
    path = get_config_path()
    if not os.path.exists(path):
        cfg = {
            "companies": {},
            "sites": DEFAULT_SITES.copy(),
            "rates": {"client": {}, "daily": {}, "overtime": 4},
            "all_workers": [],
            "assignments": {site: [] for site in DEFAULT_SITES},
            # Track which months have been added per site (YYYY-MM strings)
            "timesheets_months": {site: ["2025-07"] for site in DEFAULT_SITES}
        }
        with open(path, 'w') as f:
            json.dump(cfg, f, indent=4)
        return cfg
    with open(path, 'r') as f:
        cfg = json.load(f)
    changed = False
    # Ensure timesheets_months exists and each site has at least July 2025
    if 'timesheets_months' not in cfg:
        cfg['timesheets_months'] = {site: ["2025-07"] for site in DEFAULT_SITES}
        changed = True
    else:
        for site in DEFAULT_SITES:
            if site not in cfg['timesheets_months']:
                cfg['timesheets_months'][site] = ["2025-07"]
                changed = True
    # ... rest of key ensures ...
    if changed:
        with open(path, 'w') as f:
            json.dump(cfg, f, indent=4)
    return cfg
    with open(path, 'r') as f:
        cfg = json.load(f)
    changed = False
    # Ensure all keys exist
    if 'all_workers' not in cfg:
        cfg['all_workers'] = []
        changed = True
    if 'assignments' not in cfg:
        cfg['assignments'] = {site: [] for site in DEFAULT_SITES}
        changed = True
    if 'sites' not in cfg:
        cfg['sites'] = DEFAULT_SITES.copy()
        changed = True
    # Sync sites and assignments
    for site in DEFAULT_SITES:
        if site not in cfg['sites']:
            cfg['sites'][site] = DEFAULT_SITES[site]
            changed = True
        if site not in cfg['assignments']:
            cfg['assignments'][site] = []
            changed = True
    # Migrate legacy
    if 'workers' in cfg:
        for site, lst in cfg['workers'].items():
            for e in lst:
                wid = e['id'] if isinstance(e, dict) else e[0]
                name = e['name'] if isinstance(e, dict) else e[1] if len(e) > 1 else ''
                if not any(w['id']==wid for w in cfg['all_workers']):
                    cfg['all_workers'].append({'id': wid, 'name': name})
                if wid not in cfg['assignments'].get(site, []):
                    cfg['assignments'][site].append(wid)
        del cfg['workers']
        changed = True
    if changed:
        with open(path, 'w') as f:
            json.dump(cfg, f, indent=4)
    return cfg

def save_config(cfg):
    with open(get_config_path(), 'w') as f:
        json.dump(cfg, f, indent=4)

class WorkersPage(QtWidgets.QWidget):
    backRequested = QtCore.pyqtSignal()
    workerListChanged = QtCore.pyqtSignal()

    def __init__(self, config, parent=None):
        super().__init__(parent)
        self.config = config
        self.init_ui()

    def init_ui(self):
        layout = QtWidgets.QVBoxLayout(self)

        # Back button (top left)
        back = QtWidgets.QPushButton("Back")
        back.clicked.connect(self.backRequested)
        layout.addWidget(back, alignment=QtCore.Qt.AlignLeft)

        # --- Modern Button Row ---
        btn_row_bg = QtWidgets.QWidget()
        btn_row_bg.setStyleSheet("background: #f5f5f5; border-radius: 18px;")
        btn_row_layout = QtWidgets.QHBoxLayout(btn_row_bg)
        btn_row_layout.setContentsMargins(24, 12, 24, 12)
        self.add_btn = QtWidgets.QPushButton("Add Worker")
        self.edit_btn = QtWidgets.QPushButton("Edit Worker")
        self.delete_btn = QtWidgets.QPushButton("Delete Worker")

        # Make buttons bigger and apply modern style
        for btn in (self.add_btn, self.edit_btn, self.delete_btn):
            btn.setFixedHeight(50)
            btn.setMinimumWidth(140)
            btn.setStyleSheet("""
                QPushButton {
                    font-size: 16px;
                    border-radius: 16px;
                    padding: 10px 24px;
                    background-color: #4682f0;
                    color: white;

                }
                QPushButton:hover {
                    background-color: #2456a7;
                }
            """)
        btn_row_layout.addStretch()
        btn_row_layout.addWidget(self.add_btn)
        btn_row_layout.addSpacing(16)
        btn_row_layout.addWidget(self.edit_btn)
        btn_row_layout.addSpacing(16)
        btn_row_layout.addWidget(self.delete_btn)
        btn_row_layout.addStretch()
        layout.addWidget(btn_row_bg)
        layout.addSpacing(8)

        # Connect buttons
        self.add_btn.clicked.connect(self.add_worker_popup)
        self.edit_btn.clicked.connect(self.edit_worker)
        self.delete_btn.clicked.connect(self.delete_worker)

        # List Widgets for JKGC and JKGE in light grey groupboxes
        lists_layout = QtWidgets.QHBoxLayout()
        self.jkgc_list = QtWidgets.QListWidget(); self.jkgc_list.setMinimumWidth(230)
        self.jkge_list = QtWidgets.QListWidget(); self.jkge_list.setMinimumWidth(230)
        lists_layout.addWidget(self._groupbox('JKGC', self.jkgc_list))
        lists_layout.addWidget(self._groupbox('JKGE', self.jkge_list))
        layout.addLayout(lists_layout)

        self.setLayout(layout)
        self.refresh_all()

    def _groupbox(self, label, widget):
        gb = QtWidgets.QGroupBox(label)
        gb.setStyleSheet("""
            QGroupBox {
                background: #f5f5f5;
                border: 2px solid #e0e0e0;
                border-radius: 18px;
                font-size: 13px;
                
                margin-top: 12px;
            }
            QGroupBox:title {
                subcontrol-origin: margin;
                left: 18px;
                padding: 0 6px 0 6px;
            }
        """)
        v = QtWidgets.QVBoxLayout(gb)
        v.addWidget(widget)
        return gb

    def refresh_all(self):
        self.jkgc_list.clear()
        self.jkge_list.clear()
        def sort_key(w):
            prefix = w['id'][:4].upper()
            num = int(w['id'][4:]) if w['id'][4:].isdigit() else 0
            return (prefix, num)
        workers = sorted(self.config['all_workers'], key=sort_key)
        for w in workers:
            wid = w['id'].upper()
            entry = f"{wid} - {w['name'].upper()}"
            if wid.startswith('JKGC'):
                self.jkgc_list.addItem(entry)
            elif wid.startswith('JKGE'):
                self.jkge_list.addItem(entry)

    def add_worker_popup(self):
        dlg = QtWidgets.QDialog(self)
        dlg.setWindowTitle("Add Worker")
        f = QtWidgets.QFormLayout(dlg)
        id_box = QtWidgets.QLineEdit()
        name_box = QtWidgets.QLineEdit()
        id_box.textChanged.connect(lambda t: id_box.setText(t.upper()))
        name_box.textChanged.connect(lambda t: name_box.setText(t.upper()))
        f.addRow("Worker ID:", id_box)
        f.addRow("Name:", name_box)
        btns = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Save | QtWidgets.QDialogButtonBox.Cancel)
        f.addRow(btns)
        def save():
            wid = id_box.text().strip().upper()
            name = name_box.text().strip().upper()
            if not wid or not name:
                QtWidgets.QMessageBox.warning(dlg, "Error", "Please enter ID and Name."); return
            for w in self.config['all_workers']:
                if w['id'].upper() == wid:
                    QtWidgets.QMessageBox.warning(dlg, "Error", "Worker ID already exists!"); return
            self.config['all_workers'].append({'id': wid, 'name': name})
            save_config(self.config)
            self.refresh_all()
            self.workerListChanged.emit()
            dlg.accept()
        btns.accepted.connect(save)
        btns.rejected.connect(dlg.reject)
        dlg.exec_()

    def edit_worker(self):
        item = self.jkgc_list.currentItem() or self.jkge_list.currentItem()
        if not item:
            QtWidgets.QMessageBox.warning(self, "Error", "Select a worker to edit."); return
        wid, name = item.text().split(' - ', 1)
        dlg = QtWidgets.QDialog(self)
        dlg.setWindowTitle("Edit Worker")
        f = QtWidgets.QFormLayout(dlg)
        id_box = QtWidgets.QLineEdit(wid); id_box.setReadOnly(True)
        name_box = QtWidgets.QLineEdit(name)
        name_box.textChanged.connect(lambda t: name_box.setText(t.upper()))
        f.addRow("Worker ID:", id_box)
        f.addRow("Name:", name_box)
        btns = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Save | QtWidgets.QDialogButtonBox.Cancel)
        f.addRow(btns)
        def save():
            new_name = name_box.text().strip().upper()
            for w in self.config['all_workers']:
                if w['id'].upper() == wid:
                    w['name'] = new_name
            save_config(self.config)
            self.refresh_all()
            self.workerListChanged.emit()
            dlg.accept()
        btns.accepted.connect(save)
        btns.rejected.connect(dlg.reject)
        dlg.exec_()

    def delete_worker(self):
        item = self.jkgc_list.currentItem() or self.jkge_list.currentItem()
        if not item:
            QtWidgets.QMessageBox.warning(self, "Error", "Select a worker to delete."); return
        wid = item.text().split(' - ')[0]
        confirm = QtWidgets.QMessageBox.question(self, "Confirm Delete",
            f"Are you sure you want to delete worker {wid}? This will remove all assignments.",
            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No)
        if confirm != QtWidgets.QMessageBox.Yes: return
        self.config['all_workers'] = [w for w in self.config['all_workers'] if w['id'].upper() != wid]
        for site, lst in self.config.get('assignments', {}).items():
            self.config['assignments'][site] = [
                e for e in lst if not ((isinstance(e, dict) and e['id'].upper() == wid) or (isinstance(e, str) and e.upper() == wid))
            ]
        save_config(self.config)
        self.refresh_all()
        self.workerListChanged.emit()


class AssignPage(QtWidgets.QWidget):
    backRequested = QtCore.pyqtSignal()
    def __init__(self, config, parent=None):
        super().__init__(parent); self.config = config; self.init_ui()
    def init_ui(self):
        l = QtWidgets.QVBoxLayout(self)
        b = QtWidgets.QPushButton("Back"); b.clicked.connect(self.backRequested); l.addWidget(b,alignment=QtCore.Qt.AlignLeft)
        hl = QtWidgets.QHBoxLayout(); hl.addWidget(QtWidgets.QLabel("Site:"))
        self.site_cb=QtWidgets.QComboBox(); self.site_cb.addItem(""); self.site_cb.addItems(sorted(self.config['sites']));
        self.site_cb.currentTextChanged.connect(self.refresh_lists); hl.addWidget(self.site_cb); hl.addStretch(); l.addLayout(hl)
        l.addWidget(QtWidgets.QLabel("Assigned:"))
        self.assigned_list=QtWidgets.QListWidget(); l.addWidget(self.assigned_list)
        ctr=QtWidgets.QHBoxLayout(); ctr.addWidget(QtWidgets.QLabel("Skill:")); 
        self.skill_sel=QtWidgets.QComboBox(); self.skill_sel.addItems(SKILLS)
        ctr.addWidget(self.skill_sel)
        self.add_btn=QtWidgets.QPushButton("Add"); self.remove_btn=QtWidgets.QPushButton("Remove"); self.edit_skill_btn=QtWidgets.QPushButton("Edit Skill")
        self.add_btn.clicked.connect(self.add_assign)
        self.remove_btn.clicked.connect(self.remove_assign)
        self.edit_skill_btn.clicked.connect(self.edit_skill)
        ctr.addWidget(self.add_btn); ctr.addWidget(self.remove_btn); ctr.addWidget(self.edit_skill_btn)
        ctr.addStretch(); l.addLayout(ctr)
        l.addWidget(QtWidgets.QLabel("Unassigned:"))
        pl=QtWidgets.QHBoxLayout();
        gc=QtWidgets.QVBoxLayout(); gc.addWidget(QtWidgets.QLabel("JKGC")); self.pool_gc=QtWidgets.QListWidget(); gc.addWidget(self.pool_gc); pl.addLayout(gc)
        ge=QtWidgets.QVBoxLayout(); ge.addWidget(QtWidgets.QLabel("JKGE")); self.pool_ge=QtWidgets.QListWidget(); ge.addWidget(self.pool_ge); pl.addLayout(ge)
        l.addLayout(pl); self.setLayout(l)
        self.refresh_lists()

    def refresh_all(self):
        self.site_cb.clear()
        self.site_cb.addItem("")
        self.site_cb.addItems(sorted(self.config['sites']))
        self.refresh_lists()

    def refresh_lists(self):
        site=self.site_cb.currentText(); self.assigned_list.clear(); self.pool_gc.clear(); self.pool_ge.clear()
        if not site:
            for w in self.config['all_workers']:
                item=f"{w['id']} - {w['name']}"
                if w['id'].startswith('JKGC'): self.pool_gc.addItem(item)
                elif w['id'].startswith('JKGE'): self.pool_ge.addItem(item)
            return
        # assigned
        for e in self.config['assignments'].get(site,[]):
            wid, skl = (e['id'], e.get('skill',DEFAULT_SKILL)) if isinstance(e,dict) else (e,DEFAULT_SKILL)
            w=next((x for x in self.config['all_workers'] if x['id']==wid),None)
            if w: self.assigned_list.addItem(f"{wid} - {w['name']} ({skl})")
        # unassigned
        assigned_ids=[e['id'] if isinstance(e,dict) else e for e in self.config['assignments'].get(site,[])]
        for w in self.config['all_workers']:
            if w['id'] not in assigned_ids:
                item=f"{w['id']} - {w['name']}"
                if w['id'].startswith('JKGC'): self.pool_gc.addItem(item)
                elif w['id'].startswith('JKGE'): self.pool_ge.addItem(item)

    def add_assign(self):
        site=self.site_cb.currentText(); item=self.pool_gc.currentItem() or self.pool_ge.currentItem()
        if not site or not item: return
        wid=item.text().split(' - ')[0]; skl=self.skill_sel.currentText()
        self.config['assignments'].setdefault(site,[]).append({'id':wid,'skill':skl})
        save_config(self.config); self.refresh_lists()

    def remove_assign(self):
        site=self.site_cb.currentText(); item=self.assigned_list.currentItem()
        if not site or not item: return
        wid=item.text().split(' - ')[0]; lst=self.config['assignments'].get(site,[])
        for e in lst:
            if (isinstance(e,dict) and e['id']==wid) or (isinstance(e,str) and e==wid): lst.remove(e); break
        save_config(self.config); self.refresh_lists()

    def edit_skill(self):
        site = self.site_cb.currentText()
        item = self.assigned_list.currentItem()
        if not site or not item:
            QtWidgets.QMessageBox.warning(self, "Error", "Select a worker to edit their skill."); return
        txt = item.text()
        wid = txt.split(' - ')[0]
        # Find the assignment entry
        assignments = self.config['assignments'].get(site, [])
        entry = next((e for e in assignments if (isinstance(e, dict) and e['id'] == wid) or (isinstance(e, str) and e == wid)), None)
        if entry is None:
            QtWidgets.QMessageBox.warning(self, "Error", "Could not find assignment."); return
        current_skill = entry.get('skill', DEFAULT_SKILL) if isinstance(entry, dict) else DEFAULT_SKILL

        # Show pop-up to change skill
        dlg = QtWidgets.QDialog(self)
        dlg.setWindowTitle("Edit Skill")
        f = QtWidgets.QFormLayout(dlg)
        skill_cb = QtWidgets.QComboBox()
        skill_cb.addItems(SKILLS)
        skill_cb.setCurrentText(current_skill)
        f.addRow("Skill:", skill_cb)
        btns = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Save | QtWidgets.QDialogButtonBox.Cancel)
        f.addRow(btns)
        def save():
            new_skill = skill_cb.currentText()
            if isinstance(entry, dict):
                entry['skill'] = new_skill
            else:
                # Convert string to dict if needed
                idx = assignments.index(entry)
                assignments[idx] = {'id': wid, 'skill': new_skill}
            save_config(self.config)
            self.refresh_lists()
            dlg.accept()
        btns.accepted.connect(save)
        btns.rejected.connect(dlg.reject)
        dlg.exec_()



import os
import json
import calendar
import pandas as pd
from datetime import datetime, timedelta
from PyQt5 import QtWidgets, QtCore
from openpyxl import load_workbook

# Default values
DEFAULT_SKILL = "GENERAL WORKER"
SKILL_BASE_RATES = {
    "GENERAL WORKER": 25,
    "RIGGER/SIGNALMAN": 26,
    "LIFTING SUPERVISOR": 28,
    "WELDER": 30,
}
default_rate_col_offset = 3

class TimesheetPage(QtWidgets.QWidget):
    SUMMARY_COLS = 6

    def __init__(self, config, site, year, month, parent=None):
        super().__init__(parent)
        self.config = config
        self.site = site
        self.year = year
        self.month = month
        self.current_workers = []
        self.current_row = None
        self.current_col = None
        self.init_ui()

    def init_ui(self):
        layout = QtWidgets.QVBoxLayout(self)
        self.tbl = QtWidgets.QTableWidget()
        self.tbl.cellClicked.connect(self.on_cell_clicked)
        self.tbl.cellChanged.connect(self.on_rate_or_hour_changed)
        self.tbl.installEventFilter(self)
        layout.addWidget(self.tbl)

        ctrl = QtWidgets.QHBoxLayout()
        times = [
            "0700","0730","0800","0830","0900","1700","1800","1900",
            "1930","2000","2030","2100","2130","2200","2230","2300",
            "2330","0000"
        ]
        self.start_cb = QtWidgets.QComboBox(); self.start_cb.addItems(times)
        self.start_cb.setCurrentText('0800')
        self.end_cb   = QtWidgets.QComboBox(); self.end_cb.addItems(times)
        self.lchk     = QtWidgets.QCheckBox("LTW")
        self.ok_btn   = QtWidgets.QPushButton("OK"); self.ok_btn.clicked.connect(self.on_ok)
        self.export_btn = QtWidgets.QPushButton("Export to Excel"); self.export_btn.clicked.connect(self.on_export)
        self.sync_btn   = QtWidgets.QPushButton("Sync to OneDrive"); self.sync_btn.clicked.connect(self.on_sync)

        self.ok_btn.setShortcut(QtCore.Qt.Key_Return)
        self.ok_btn.setShortcut(QtCore.Qt.Key_Enter)

        ctrl.addWidget(QtWidgets.QLabel("Start:")); ctrl.addWidget(self.start_cb)
        ctrl.addWidget(QtWidgets.QLabel("End:"));   ctrl.addWidget(self.end_cb)
        ctrl.addWidget(self.lchk);                  ctrl.addWidget(self.ok_btn)
        ctrl.addStretch()
        ctrl.addWidget(self.export_btn)
        ctrl.addWidget(self.sync_btn)
        layout.addLayout(ctrl)

        self.load()

    # ... (load, save, eventFilter, on_cell_clicked, on_ok, on_rate_or_hour_changed, update_summary identical to previous version) ...

    def on_export(self):
        # Build DataFrame
        rows, cols = [], self.tbl.columnCount()
        headers = [self.tbl.horizontalHeaderItem(c).text() for c in range(cols)]
        for r in range(self.tbl.rowCount()):
            row_data = {headers[c]: self.tbl.item(r,c).text() if self.tbl.item(r,c) else ''
                        for c in range(cols)}
            row_data['Worker'] = self.tbl.verticalHeaderItem(r).text()
            rows.append(row_data)
        df = pd.DataFrame(rows)[['Worker'] + headers]
        fname = f"timesheet_{self.site}_{self.year}_{self.month}.xlsx"
        df.to_excel(fname, index=False)
        QtWidgets.QMessageBox.information(self, "Exported", f"Saved local Excel: {fname}")

    def on_sync(self):
        # Prompt for OneDrive workbook
        od_path, _ = QtWidgets.QFileDialog.getOpenFileName(
            self, "Select OneDrive Timesheet Workbook", os.path.expanduser("~"),
            "Excel Files (*.xlsx *.xlsm)")
        if not od_path:
            return
        # Load existing workbook and add/overwrite a sheet
        sheet_name = f"{self.site}_{self.year}_{self.month}"  # e.g. SiteA_2025_07
        try:
            wb = load_workbook(od_path)
            if sheet_name in wb.sheetnames:
                wb.remove(wb[sheet_name])
            ws = wb.create_sheet(title=sheet_name)
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Error", f"Failed to open workbook:\n{e}")
            return
        # Gather table data
        headers = [self.tbl.horizontalHeaderItem(c).text() for c in range(self.tbl.columnCount())]
        # Write header row
        for ci, h in enumerate(['Worker'] + headers, start=1):
            ws.cell(row=1, column=ci, value=h)
        # Write each worker row
        for r in range(self.tbl.rowCount()):
            worker_text = self.tbl.verticalHeaderItem(r).text()
            ws.cell(row=r+2, column=1, value=worker_text)
            for c, h in enumerate(headers, start=2):
                val = self.tbl.item(r, c-2).text() if self.tbl.item(r, c-2) else ''
                ws.cell(row=r+2, column=c, value=val)
        # Save workbook back to OneDrive
        try:
            wb.save(od_path)
            QtWidgets.QMessageBox.information(
                self, "Synced", f"Timesheet synced to '{sheet_name}' in:\n{od_path}")
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Error", f"Failed to save workbook:\n{e}")

# Site entry page with month tabs
class SitesPage(QtWidgets.QWidget):
    backRequested = QtCore.pyqtSignal()

    def __init__(self, config, parent=None):
        super().__init__(parent)
        self.config = config
        self.init_ui()

    def init_ui(self):
        v = QtWidgets.QVBoxLayout(self)
        b = QtWidgets.QPushButton("Back"); b.clicked.connect(self.backRequested)
        v.addWidget(b, alignment=QtCore.Qt.AlignLeft)
        hs = QtWidgets.QHBoxLayout(); hs.addWidget(QtWidgets.QLabel("Site:"))
        self.site_cb = QtWidgets.QComboBox(); self.site_cb.addItem("")
        self.site_cb.addItems(sorted(self.config['sites']))
        self.site_cb.currentTextChanged.connect(self.build_tabs)
        hs.addWidget(self.site_cb); hs.addStretch(); v.addLayout(hs)
        self.tabs = QtWidgets.QTabWidget(); v.addWidget(self.tabs)
        self.build_tabs()

    def build_tabs(self):
        self.tabs.clear()
        site = self.site_cb.currentText()
        if not site:
            return
        months = self.config.setdefault('timesheets_months', {}).setdefault(site, [])
        for i, m in enumerate(months):
            y, mo = map(int, m.split('-'))
            label = datetime(y, mo, 1).strftime('%b %Y')
            page = TimesheetPage(self.config, site, y, mo)
            idx = self.tabs.addTab(page, label)
            # Important: use functools.partial or default arg to capture month correctly
            del_btn = QtWidgets.QPushButton('Delete')
            del_btn.clicked.connect(lambda _, m=m: self.delete_month(m))
            self.tabs.tabBar().setTabButton(idx, QtWidgets.QTabBar.RightSide, del_btn)
        btn = QtWidgets.QPushButton('Add Month')
        idx = self.tabs.addTab(QtWidgets.QWidget(), "")
        self.tabs.tabBar().setTabButton(idx, QtWidgets.QTabBar.RightSide, btn)
        btn.clicked.connect(self.add_month)

    def add_month(self):
        site = self.site_cb.currentText()
        if not site: return
        lst = self.config['timesheets_months'].setdefault(site, ['2025-07'])
        last = lst[-1]
        y, mo = map(int, last.split('-'))
        if mo == 12: y += 1; mo = 1
        else: mo += 1
        lst.append(f"{y:04d}-{mo:02d}")
        save_config(self.config)
        self.build_tabs()

    def delete_month(self, month):
        site = self.site_cb.currentText()
        if not site: return
        confirm = QtWidgets.QMessageBox.question(self, "Confirm Delete",
            f"Are you sure you want to delete the month {month} for site {site}?",
            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No)
        if confirm != QtWidgets.QMessageBox.Yes:
            return
        months = self.config['timesheets_months'][site]
        if month in months:
            months.remove(month)
            save_config(self.config)
            self.build_tabs()

# Main Window with Logo Background, 2-row Buttons
from PyQt5 import QtWidgets, QtGui, QtCore
import sys

class LogoWidget(QtWidgets.QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.logo = QtGui.QPixmap("JKG LOGO.jpg")
        self.setMinimumSize(900, 600)

    def paintEvent(self, event):
        painter = QtGui.QPainter(self)
        painter.setRenderHint(QtGui.QPainter.SmoothPixmapTransform)
        # Draw logo centered and scaled
        if not self.logo.isNull():
            rect = self.rect()
            logo_scaled = self.logo.scaled(
                rect.width()//2, rect.height()//2,
                QtCore.Qt.KeepAspectRatio,
                QtCore.Qt.SmoothTransformation)
            x = (rect.width() - logo_scaled.width()) // 2
            y = (rect.height() - logo_scaled.height()) // 2
            painter.drawPixmap(x, y, logo_scaled)

from PyQt5 import QtWidgets, QtGui, QtCore
import sys

class LogoWidget(QtWidgets.QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.logo = QtGui.QPixmap("JKG LOGO.jpg")
        self.setMinimumSize(900, 600)

    def paintEvent(self, event):
        painter = QtGui.QPainter(self)
        painter.setRenderHint(QtGui.QPainter.SmoothPixmapTransform)
        # Draw logo centered and scaled to fit window
        if not self.logo.isNull():
            rect = self.rect()
            logo_scaled = self.logo.scaled(
                rect.size(), QtCore.Qt.KeepAspectRatio, QtCore.Qt.SmoothTransformation)
            x = (rect.width() - logo_scaled.width()) // 2
            y = (rect.height() - logo_scaled.height()) // 2
            painter.drawPixmap(x, y, logo_scaled)

class HomePage(QtWidgets.QWidget):
    def __init__(self, btn_texts, btn_slots):
        super().__init__()
        self.logo_widget = LogoWidget(self)
        self.logo_widget.lower()
        layout = QtWidgets.QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)
        # App title
        title = QtWidgets.QLabel("Welcome to the JKG App")
        title.setAlignment(QtCore.Qt.AlignCenter)
        title.setFont(QtGui.QFont("Segoe UI", 26, QtGui.QFont.Bold))
        title.setStyleSheet("color: #035397; margin: 40px 0 10px 0; text-shadow: 0px 1px 4px #FFF;")
        layout.addWidget(title)
        layout.addStretch()
        # Grid buttons
        grid = QtWidgets.QGridLayout()
        grid.setHorizontalSpacing(10)
        grid.setVerticalSpacing(10)
        for i, (txt, slot) in enumerate(zip(btn_texts, btn_slots)):
            btn = QtWidgets.QPushButton(txt)
            btn.setFixedSize(150, 70)
            btn.setStyleSheet(
                "font-size:20px;background:rgba(0,122,204,0.78);color:white;border-radius:15px;"
                "border: 2px solid #007ACC;"
            )
            btn.clicked.connect(slot)
            row, col = divmod(i, 2)
            grid.addWidget(btn, row, col, alignment=QtCore.Qt.AlignCenter)
        grid.setAlignment(QtCore.Qt.AlignCenter)
        layout.addLayout(grid)
        layout.addStretch()

    def resizeEvent(self, event):
        self.logo_widget.setGeometry(self.rect())
        super().resizeEvent(event)

from PyQt5 import QtWidgets, QtGui, QtCore
import sys

class LogoWidget(QtWidgets.QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.logo = QtGui.QPixmap("JKG LOGO.jpg")
        self.setMinimumSize(900, 600)

    def paintEvent(self, event):
        painter = QtGui.QPainter(self)
        painter.setRenderHint(QtGui.QPainter.SmoothPixmapTransform)
        if not self.logo.isNull():
            rect = self.rect()
            # Always scale up, but not more than window (logo always sharp)
            logo_scaled = self.logo.scaled(
                rect.size(), QtCore.Qt.KeepAspectRatioByExpanding, QtCore.Qt.SmoothTransformation)
            x = (rect.width() - logo_scaled.width()) // 2
            y = (rect.height() - logo_scaled.height()) // 2
            # Blur effect
            blur_radius = 32  # Change this for more/less blur
            image = logo_scaled.toImage()
            for _ in range(blur_radius // 4):  # Quick-and-dirty box blur (pseudo)
                image = image.scaled(image.width() // 2, image.height() // 2, QtCore.Qt.IgnoreAspectRatio, QtCore.Qt.SmoothTransformation)
                image = image.scaled(image.width() * 2, image.height() * 2, QtCore.Qt.IgnoreAspectRatio, QtCore.Qt.SmoothTransformation)
            painter.drawImage(x, y, image)

class HomePage(QtWidgets.QWidget):
    def __init__(self, btn_texts, btn_slots):
        super().__init__()
        self.logo_widget = LogoWidget(self)
        self.logo_widget.lower()
        layout = QtWidgets.QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)
        layout.addStretch()
        # Grid buttons
        grid = QtWidgets.QGridLayout()
        grid.setHorizontalSpacing(10)
        grid.setVerticalSpacing(10)
        for i, (txt, slot) in enumerate(zip(btn_texts, btn_slots)):
            btn = QtWidgets.QPushButton(txt)
            btn.setFixedSize(150, 70)
            btn.setStyleSheet(
                "font-size:20px;background:rgba(0,122,204,0.78);color:white;border-radius:15px;"
                "border: 2px solid #007ACC;"
            )
            btn.clicked.connect(slot)
            row, col = divmod(i, 2)
            grid.addWidget(btn, row, col, alignment=QtCore.Qt.AlignCenter)
        grid.setAlignment(QtCore.Qt.AlignCenter)
        layout.addLayout(grid)
        # App title below logo but above buttons
        title = QtWidgets.QLabel("Welcome to the JKG App")
        title.setAlignment(QtCore.Qt.AlignCenter)
        title.setFont(QtGui.QFont("Segoe UI", 26, QtGui.QFont.Bold))
        title.setStyleSheet("color: #035397; margin: 0 0 25px 0;")
        layout.insertWidget(layout.count()-1, title)
        layout.addStretch()

    def resizeEvent(self, event):
        self.logo_widget.setGeometry(self.rect())
        super().resizeEvent(event)

from PyQt5 import QtWidgets, QtGui, QtCore
import sys

class LogoWidget(QtWidgets.QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.logo = QtGui.QPixmap("JKG LOGO.jpg")
        self.setMinimumSize(900, 600)
        self.setAttribute(QtCore.Qt.WA_StyledBackground, True)
        self.setStyleSheet("background-color: white;")

    def paintEvent(self, event):
        painter = QtGui.QPainter(self)
        painter.setRenderHint(QtGui.QPainter.SmoothPixmapTransform)
        if not self.logo.isNull():
            rect = self.rect()
            # Increase logo size (e.g., to 70% of window size)
            w = int(rect.width() * 0.7)
            h = int(rect.height() * 0.7)
            logo_scaled = self.logo.scaled(w, h, QtCore.Qt.KeepAspectRatio, QtCore.Qt.SmoothTransformation)
            image = logo_scaled.toImage()
            for _ in range(8):
                image = image.scaled(image.width() // 2, image.height() // 2, QtCore.Qt.IgnoreAspectRatio, QtCore.Qt.SmoothTransformation)
                image = image.scaled(image.width() * 2, image.height() * 2, QtCore.Qt.IgnoreAspectRatio, QtCore.Qt.SmoothTransformation)
            x = (rect.width() - logo_scaled.width()) // 2
            y = (rect.height() - logo_scaled.height()) // 2
            painter.drawImage(x, y, image)

class HomePage(QtWidgets.QWidget):
    def __init__(self, btn_texts, btn_slots):
        super().__init__()
        self.setAttribute(QtCore.Qt.WA_StyledBackground, True)
        self.setStyleSheet("background-color: white;")
        self.logo_widget = LogoWidget(self)
        self.logo_widget.lower()
        layout = QtWidgets.QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)
        layout.addStretch(1)
        # Title now very close to the logo
        title = QtWidgets.QLabel("Welcome to the JKG Application")
        title.setAlignment(QtCore.Qt.AlignHCenter | QtCore.Qt.AlignBottom)
        title.setFont(QtGui.QFont("Segoe UI", 30, QtGui.QFont.Bold))
        title.setStyleSheet("color: #17418a; margin-top: 0px; margin-bottom: 6px;")
        layout.addWidget(title, alignment=QtCore.Qt.AlignHCenter)
        # Button grid, centered under logo
        grid = QtWidgets.QGridLayout()
        grid.setHorizontalSpacing(16)
        grid.setVerticalSpacing(16)
        button_color = "#e3f0fb"  # Light blue
        border_color = "#b3d1ee"
        text_color = "#145494"
        for i, (txt, slot) in enumerate(zip(btn_texts, btn_slots)):
            btn = QtWidgets.QPushButton(txt)
            btn.setFixedSize(170, 72)
            btn.setStyleSheet(
                f"font-size:20px;background:{button_color};color:{text_color};border-radius:15px;"
                f"border: 2px solid {border_color}; font-weight:600;"
            )
            btn.clicked.connect(slot)
            row, col = divmod(i, 2)
            grid.addWidget(btn, row, col, alignment=QtCore.Qt.AlignCenter)
        grid.setAlignment(QtCore.Qt.AlignHCenter)
        layout.addLayout(grid)
        layout.addStretch(2)

    def resizeEvent(self, event):
        self.logo_widget.setGeometry(self.rect())
        super().resizeEvent(event)

class MainWindow(QtWidgets.QMainWindow):
    def __init__(self):
        super().__init__()
        self.config = load_config()
        self.setWindowTitle("Welcome to the JKG Application")
        self.resize(900, 600)
        self.setAttribute(QtCore.Qt.WA_StyledBackground, True)
        self.setStyleSheet("background-color: white;")
        self.init_ui()

    def init_ui(self):
        btn_texts = ["WORKERS", "ASSIGN", "SITES"]
        btn_slots = [self.show_workers, self.show_assign, self.show_sites]
        self.home = HomePage(btn_texts, btn_slots)
        self.wp = WorkersPage(self.config)
        self.wp.backRequested.connect(lambda: self.stack.setCurrentWidget(self.home))
        self.ap = AssignPage(self.config)
        self.ap.backRequested.connect(lambda: self.stack.setCurrentWidget(self.home))
        self.sp = SitesPage(self.config)
        self.sp.backRequested.connect(lambda: self.stack.setCurrentWidget(self.home))
        self.stack = QtWidgets.QStackedWidget()
        self.stack.setStyleSheet("background-color: white;")
        self.stack.addWidget(self.home)
        self.stack.addWidget(self.wp)
        self.stack.addWidget(self.ap)
        self.stack.addWidget(self.sp)
        self.setCentralWidget(self.stack)

    def show_workers(self):
        self.stack.setCurrentWidget(self.wp)
    def show_assign(self):
        self.stack.setCurrentWidget(self.ap)
    def show_sites(self):
        self.stack.setCurrentWidget(self.sp)

if __name__ == '__main__':
    app = QtWidgets.QApplication(sys.argv)
    app.setStyleSheet("QWidget { background-color: white; }")
    win = MainWindow()
    win.show()
    sys.exit(app.exec_())
